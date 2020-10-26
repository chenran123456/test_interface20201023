from openpyxl import load_workbook
from conf import read_path
import os

class DoExcel:
    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name

    def do_excel(self):
        wb=load_workbook(self.file_path)
        sheet=wb[self.sheet_name]
        test_data=[] #所有行的数据存到这个列表里面
        for i in range(2,sheet.max_row+1):
            sub_data={}  #每一行用例的数据存在一个字典里面
            sub_data['case_id']=sheet.cell(i,1).value
            sub_data['title']=sheet.cell(i,2).value
            sub_data['method']=sheet.cell(i,3).value
            sub_data['url']=sheet.cell(i,4).value
            sub_data['param']=sheet.cell(i,5).value
            test_data.append(sub_data)
        return test_data

if __name__ == '__main__':
     path=os.path.join(read_path.test_data_path)
     test_data=DoExcel(path,'test_data').do_excel()
     print(test_data)




